from __future__ import annotations

import copy
import csv
import json
import os
import shlex
import subprocess
from time import perf_counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from threading import BoundedSemaphore, RLock
from typing import Any, Callable, Dict, List, Optional, Sequence

from impl.core.config import get_runtime_config
from impl.core.schema import (
    MockCasesResponse,
    MockDatasetsResponse,
    RunChainResponse,
    normalize_attribute_result,
    normalize_case_pool_table,
    normalize_check_report,
    normalize_cluster_summary,
    normalize_frontend_view,
    normalize_judge_result,
    normalize_run_trace,
    normalize_trace_table_row,
    to_dict,
)
from impl.core.mock_agent import load_live_schema
from impl.core.schema.fixture import load_fixture

PROJECT_IDS = [
    "QA",
    "client_search",
    "marketting-planning",
    "marketting-planning-intent",
]
PROJECT_ID = PROJECT_IDS[0]

_PROJECT_FLOW_CACHE: Dict[str, Dict[str, Any]] = {}
_MOCK_CASE_CACHE: Dict[str, Dict[str, Any]] = {}
_PROJECT_LOCKS: Dict[str, RLock] = {}
_CACHE_LOCK = RLock()
_PROGRESS_LOCK = RLock()
_PROGRESS_DONE = 0
_SHOW_FLOW_PROGRESS = False
API_CALL_TIMEOUT_SECONDS = 600.0
_RUNTIME_CONFIG = get_runtime_config()
API_BASE_URL = os.environ.get(
    "API_CHECK_BASE_URL",
    f"http://{_RUNTIME_CONFIG.uat.host}:{_RUNTIME_CONFIG.uat.port}",
).rstrip("/")
API_MAX_WORKERS = 8
LLM_API_MAX_WORKERS = 5
LLM_HEAVY_CASES = {"live_run", "run_chain", "judge", "attribute", "batch_run"}
_LLM_CALL_SEMAPHORE = BoundedSemaphore(LLM_API_MAX_WORKERS)


@dataclass(frozen=True)
class ApiCase:
    name: str
    path: str
    build_request: Callable[[str], Dict[str, Any]]
    expected_schema: str
    response_path: Sequence[str] = ()
    projects: Sequence[str] = tuple(PROJECT_IDS)
    llm_heavy: bool = False

    def request_for_project(self, project_id: str) -> Dict[str, Any]:
        return self.build_request(project_id)

    def call_api(self, request: Dict[str, Any]) -> Dict[str, Any]:
        status_code, response = call_api_raw(self.path, request)
        assert status_code == 200, json.dumps(response, ensure_ascii=False)
        return response

    def assert_response_schema(self, response: Dict[str, Any], project_id: str = "") -> None:
        assert_schema(_pick(response, self.response_path), self.expected_schema, project_id=project_id)

def fixture_trace(project_id: str) -> Dict[str, Any]:
    trace = load_fixture("impl.core.schema.trace.RunTrace", as_dict=True, project_id=project_id)
    live_result = trace.get("live_result") if isinstance(trace.get("live_result"), dict) else {}
    live_result["project_id"] = project_id
    trace["live_result"] = live_result
    return trace


def fixture_judge(project_id: str) -> Dict[str, Any]:
    return load_fixture("impl.core.schema.judge.JudgeResult", as_dict=True, project_id=project_id)


def fixture_attribute(project_id: str) -> Dict[str, Any]:
    return load_fixture("impl.core.schema.attribute.AttributeResult", as_dict=True, project_id=project_id)


def fixture_cluster(project_id: str) -> Dict[str, Any]:
    return load_fixture("impl.core.schema.cluster.ClusterSummary", as_dict=True, project_id=project_id)


def fixture_check() -> Dict[str, Any]:
    return load_fixture("impl.core.schema.check.CheckReport", as_dict=True)


def fixture_frontend_view(project_id: str) -> Dict[str, Any]:
    view = load_fixture("impl.core.schema.frontend.FrontendViewModel", as_dict=True)
    view["project_info"] = {**(view.get("project_info") or {}), "project_id": project_id}
    return view


def fixture_run_payload(project_id: str) -> Dict[str, Any]:
    flow = project_flow(project_id)
    return {
        "trace": flow["trace"],
        "judge": flow["judge"],
        "attribute": flow["attribute"],
        "cluster": flow["cluster"],
        "check": flow["check"],
        "frontend_view": flow["frontend_view"],
    }


def fixture_table_row_request(project_id: str) -> Dict[str, Any]:
    flow = project_flow(project_id)
    return {
        "project": project_id,
        "trace": flow["trace"],
        "judge": flow["judge"],
        "attribute": flow["attribute"],
        "check": flow["check"],
        "frontend_view": flow["frontend_view"],
    }


API_FIXTURE_CHECKS = [
    ApiCase(
        "analysis",
        "/api/analysis",
        lambda project_id: {"project": project_id},
        "impl.core.schema.project.ProjectAnalysis",
    ),
    ApiCase(
        "mock_cases",
        "/api/mock_cases",
        lambda project_id: {"project": project_id},
        "impl.core.schema.api.MockCasesResponse",
    ),
    ApiCase(
        "mock_datasets",
        "/api/mock_datasets",
        lambda project_id: {"project": project_id},
        "impl.core.schema.api.MockDatasetsResponse",
    ),
    ApiCase(
        "live_run",
        "/api/live_run",
        lambda project_id: {"project": project_id, "input": real_project_case(project_id)},
        "impl.core.schema.trace.RunTrace",
    ),
    ApiCase(
        "run_chain",
        "/api/run_chain",
        lambda project_id: {"project": project_id, "input": real_project_case(project_id)},
        "impl.core.schema.api.RunChainResponse",
    ),
    ApiCase(
        "judge",
        "/api/judge",
        lambda project_id: {"project": project_id, "trace": project_flow(project_id)["trace"]},
        "impl.core.schema.judge.JudgeResult",
    ),
    ApiCase(
        "attribute",
        "/api/attribute",
        lambda project_id: {"project": project_id, "trace": project_flow(project_id)["trace"], "judge": project_flow(project_id)["judge"]},
        "impl.core.schema.attribute.AttributeResult",
    ),
    ApiCase(
        "cluster",
        "/api/cluster",
        lambda project_id: {"project": project_id, "attributes": [project_flow(project_id)["attribute"]]},
        "impl.core.schema.cluster.ClusterSummary",
    ),
    ApiCase(
        "check",
        "/api/check",
        lambda project_id: {
            "project": project_id,
            "trace": project_flow(project_id)["trace"],
            "judge": project_flow(project_id)["judge"],
            "attribute": project_flow(project_id)["attribute"],
            "cluster": project_flow(project_id)["cluster"],
        },
        "impl.core.schema.check.CheckReport",
    ),
    ApiCase(
        "frontend_view",
        "/api/frontend_view",
        lambda project_id: {
            "project": project_id,
            "trace": project_flow(project_id)["trace"],
            "judge": project_flow(project_id)["judge"],
            "attribute": project_flow(project_id)["attribute"],
            "cluster": project_flow(project_id)["cluster"],
            "check": project_flow(project_id)["check"],
        },
        "impl.core.schema.frontend.FrontendViewModel",
    ),
    ApiCase(
        "batch_run",
        "/api/batch_run",
        lambda project_id: {"project": project_id, "cases": [real_project_case(project_id)], "concurrency": 1},
        "impl.core.schema.batch.BatchRunResult",
    ),
    ApiCase(
        "trace",
        "/api/trace",
        lambda project_id: {"project": project_id, "trace": project_flow(project_id)["trace"]},
        "impl.core.schema.trace.RunTrace",
    ),
    ApiCase(
        "table_row",
        "/api/table",
        lambda project_id: fixture_table_row_request(project_id),
        "impl.core.schema.table.TraceTableRow",
    ),
    ApiCase(
        "table_pool",
        "/api/table",
        lambda project_id: {"project": project_id, "runs": [fixture_run_payload(project_id)]},
        "impl.core.schema.table.CasePoolTable",
    ),
]


def project_lock(project_id: str) -> RLock:
    with _CACHE_LOCK:
        if project_id not in _PROJECT_LOCKS:
            _PROJECT_LOCKS[project_id] = RLock()
        return _PROJECT_LOCKS[project_id]


def flow_progress(message: str) -> None:
    if _SHOW_FLOW_PROGRESS:
        with _PROGRESS_LOCK:
            print(message, flush=True)


def real_project_case(project_id: str) -> Dict[str, Any]:
    lock = project_lock(project_id)
    with lock:
        if project_id not in _MOCK_CASE_CACHE:
            flow_progress(f"[case-cache-start] {project_id} /api/mock_cases")
            status_code, response = call_api_raw("/api/mock_cases", {"project": project_id})
            if status_code != 200:
                raise AssertionError(f"/api/mock_cases failed for {project_id}: {response}")
            cases = response.get("cases") or []
            if not cases:
                raise AssertionError(f"/api/mock_cases returned no cases for {project_id}")
            _MOCK_CASE_CACHE[project_id] = cases[0]
            flow_progress(f"[case-cache-done] {project_id} /api/mock_cases")
        return _MOCK_CASE_CACHE[project_id]


def project_flow(project_id: str) -> Dict[str, Any]:
    lock = project_lock(project_id)
    with lock:
        if project_id not in _PROJECT_FLOW_CACHE:
            flow_progress(f"[project-flow-start] {project_id} /api/run_chain")
            request = {"project": project_id, "input": real_project_case(project_id)}
            status_code, response = call_api_raw("/api/run_chain", request)
            if status_code != 200:
                raise AssertionError(f"/api/run_chain failed for {project_id}: {response}")
            assert_schema(response, "impl.core.schema.api.RunChainResponse", project_id=project_id)
            _PROJECT_FLOW_CACHE[project_id] = response
            flow_progress(f"[project-flow-done] {project_id} /api/run_chain")
        return copy.deepcopy(_PROJECT_FLOW_CACHE[project_id])


def api_case_matrix() -> List[tuple[ApiCase, str]]:
    return [(case, project_id) for case in API_FIXTURE_CHECKS for project_id in case.projects]


def _pick(value: Any, path: Sequence[str]) -> Any:
    for key in path:
        value = value[key]
    return value


def _assert_judge_business_shape(judge: Any, project_id: str, where: str) -> None:
    """api-check 层 judge 形状校验，委托给 LiveSchemaCheck（统一从 SchemaValidator 取）。"""
    result = normalize_judge_result(judge)
    assert result.trace_id
    live_schema = load_live_schema(project_id) if project_id else None
    checker = getattr(live_schema, "check", None) if live_schema else None
    if not checker:
        return
    if result.actual is not None:
        assert checker.output(result.actual), f"{where}.actual 不符合 {project_id} EXTRACT_OUTPUT_SHAPE: {result.actual}"
    if result.expected is not None:
        assert checker.reference(result.expected), f"{where}.expected 不符合 {project_id} EXTRACT_OUTPUT_SHAPE: {result.expected}"
    elif not getattr(result, "quality_flags", None) or "llm_call_failed" not in result.quality_flags:
        raise AssertionError(f"{where}.expected 缺失")


def _assert_nested_run_judges(value: Any, project_id: str, where: str) -> None:
    if not isinstance(value, dict):
        return
    if isinstance(value.get("judge"), dict):
        _assert_judge_business_shape(value["judge"], project_id, f"{where}.judge")
    runs = value.get("runs")
    if isinstance(runs, list):
        for index, run in enumerate(runs):
            if isinstance(run, dict) and isinstance(run.get("judge"), dict):
                _assert_judge_business_shape(run["judge"], project_id, f"{where}.runs[{index}].judge")


def assert_schema(value: Any, expected_schema: str, project_id: str = "") -> None:
    if expected_schema == "impl.core.schema.project.ProjectAnalysis":
        assert value["project_id"]
    elif expected_schema == "impl.core.schema.trace.RunTrace":
        assert normalize_run_trace(value).trace_id
    elif expected_schema == "impl.core.schema.judge.JudgeResult":
        _assert_judge_business_shape(value, project_id, "judge")
    elif expected_schema == "impl.core.schema.attribute.AttributeResult":
        assert normalize_attribute_result(value).trace_id
    elif expected_schema == "impl.core.schema.cluster.ClusterSummary":
        assert normalize_cluster_summary(value).project_id
    elif expected_schema == "impl.core.schema.check.CheckReport":
        assert isinstance(normalize_check_report(value).passed, bool)
    elif expected_schema == "impl.core.schema.frontend.FrontendViewModel":
        assert normalize_frontend_view(value).table_row is not None
    elif expected_schema == "impl.core.schema.table.TraceTableRow":
        assert normalize_trace_table_row(value).trace_id
    elif expected_schema == "impl.core.schema.table.CasePoolTable":
        assert normalize_case_pool_table(value).rows
        _assert_nested_run_judges(value, project_id, "case_pool_table")
    elif expected_schema == "impl.core.schema.batch.BatchRunResult":
        assert value.get("project_id") and isinstance(value.get("runs"), list)
        if value.get("table"):
            assert normalize_case_pool_table(value.get("table")).rows
        _assert_nested_run_judges(value, project_id, "batch_run")
    elif expected_schema == "impl.core.schema.api.MockCasesResponse":
        response = MockCasesResponse(project_id=value["project_id"], cases=value["cases"])
        assert response.project_id and response.cases
    elif expected_schema == "impl.core.schema.api.MockDatasetsResponse":
        response = MockDatasetsResponse(project_id=value["project_id"], datasets=value["datasets"])
        assert response.project_id and response.datasets
    elif expected_schema == "impl.core.schema.api.RunChainResponse":
        response = RunChainResponse(
            trace=normalize_run_trace(value.get("trace")),
            judge=normalize_judge_result(value.get("judge")),
            attribute=normalize_attribute_result(value.get("attribute")) if value.get("attribute") else None,
            cluster=normalize_cluster_summary(value.get("cluster")) if value.get("cluster") else None,
            check=normalize_check_report(value.get("check")) if value.get("check") else None,
            frontend_view=normalize_frontend_view(value.get("frontend_view")) if value.get("frontend_view") else None,
            table_row=normalize_trace_table_row(value.get("table_row")) if value.get("table_row") else None,
        )
        assert response.trace.trace_id and response.judge.trace_id
        _assert_judge_business_shape(value.get("judge"), project_id, "run_chain.judge")
    else:
        raise AssertionError(f"unknown expected schema: {expected_schema}")


def call_api_raw(path: str, request: Dict[str, Any], timeout: float = API_CALL_TIMEOUT_SECONDS) -> tuple[int, Dict[str, Any]]:
    completed = subprocess.run(
        curl_args(path, request),
        text=True,
        capture_output=True,
        timeout=timeout,
    )
    if completed.returncode != 0:
        body = {"curl_exit_code": completed.returncode, "stderr": completed.stderr, "stdout": completed.stdout}
        return 0, body
    stdout = completed.stdout
    marker = "\n__HTTP_STATUS__:"
    status_code = 0
    if marker in stdout:
        payload_text, _, status_text = stdout.rpartition(marker)
        stdout = payload_text
        try:
            status_code = int(status_text.strip())
        except ValueError:
            status_code = 0
    try:
        body = json.loads(stdout) if stdout else {}
    except json.JSONDecodeError:
        body = {"raw_text": stdout}
    return status_code, body


EXCEL_CELL_LIMIT = 30000


def json_cell(value: Any) -> str:
    return json.dumps(to_dict(value), ensure_ascii=False, sort_keys=True)


def pretty_json(value: Any) -> str:
    return json.dumps(to_dict(value), ensure_ascii=False, indent=2)


def excel_preview_cell(value: Any, *, limit: int = EXCEL_CELL_LIMIT) -> str:
    text = pretty_json(value)
    if len(text) <= limit:
        return text
    return json.dumps({
        "truncated_preview_only": True,
        "full_json_chars": len(text),
        "preview": text[: limit - 220],
    }, ensure_ascii=False)


def excel_curl_cell(command: str, script_path: str, *, limit: int = EXCEL_CELL_LIMIT) -> str:
    if len(command) <= limit:
        return command
    preview_limit = min(len(command), limit - 1000)
    while preview_limit > 0:
        cell = json.dumps({
            "too_long_for_excel_cell": True,
            "full_curl_chars": len(command),
            "curl_script_path": script_path,
            "preview": command[:preview_limit],
        }, ensure_ascii=False)
        if len(cell) <= limit:
            return cell
        preview_limit -= 1000
    return json.dumps({
        "too_long_for_excel_cell": True,
        "full_curl_chars": len(command),
        "curl_script_path": script_path,
    }, ensure_ascii=False)


def curl_args(path: str, request: Dict[str, Any]) -> List[str]:
    return [
        "curl",
        "-sS",
        "-X",
        "POST",
        f"{API_BASE_URL}{path}",
        "-H",
        "Content-Type: application/json",
        "--data-raw",
        json_cell(request),
        "--write-out",
        "\n__HTTP_STATUS__:%{http_code}",
    ]


def curl_command(path: str, request: Dict[str, Any]) -> str:
    return " ".join(shlex_quote(item) if any(char.isspace() or char in "'{}[]:,\"" for char in item) else item for item in curl_args(path, request))


def shlex_quote(value: str) -> str:
    return "'" + value.replace("'", "'\\''") + "'"


def replayable_curl(path: str, request: Dict[str, Any]) -> tuple[str, str]:
    command = curl_command(path, request)
    if len(command) > EXCEL_CELL_LIMIT:
        return command, f"curl too long for one Excel cell ({len(command)} chars); curl_script_path records the exact executable command"
    return command, "copyable visual curl: this is the same JSON request body used by the API call"


def run_api_case(case: ApiCase, project_id: str) -> Dict[str, Any]:
    request: Dict[str, Any] = {}
    response: Dict[str, Any]
    schema_check = "not_run"
    schema_error = ""
    status_code = 0
    started = perf_counter()
    flow_progress(f"[case-start] {project_id} {case.name} {case.path}")
    try:
        request = case.request_for_project(project_id)
        if case.llm_heavy or case.name in LLM_HEAVY_CASES:
            with _LLM_CALL_SEMAPHORE:
                status_code, response = call_api_raw(case.path, request)
        else:
            status_code, response = call_api_raw(case.path, request)
        if status_code == 200:
            try:
                case.assert_response_schema(response, project_id)
                schema_check = "pass"
            except Exception as exc:
                schema_check = "fail"
                schema_error = str(exc)
        else:
            schema_check = "http_error"
            schema_error = str(response.get("error") or response)
    except Exception as exc:
        response = {"error": str(exc)}
        schema_check = "setup_error"
        schema_error = str(exc)
    elapsed = perf_counter() - started
    source = "api"
    # 挂载 live_schema 校验：请求体是否符合 REQUEST_SHAPE
    live_schema_ok = _check_api_request_with_live_schema(project_id, request)
    curl, curl_note = replayable_curl(case.path, request) if request else ("", "request build failed")
    flow_progress(f"[case-done] {project_id} {case.name} source={source} http={status_code} schema={schema_check} elapsed={elapsed:.1f}s")
    return {
        "project": project_id,
        "case": case.name,
        "api": f"POST {case.path}",
        "http_status": status_code,
        "curl": curl,
        "curl_note": curl_note,
        "request_body": to_dict(request),
        "response_body": to_dict(response),
        "expected_schema": case.expected_schema,
        "checked_response_path": list(case.response_path),
        "schema_check": schema_check,
        "schema_error": schema_error,
        "source": source,
    }


def iter_api_project_rows(progress: Optional[Callable[[Dict[str, Any], int, int], None]] = None) -> List[Dict[str, Any]]:
    global _SHOW_FLOW_PROGRESS
    rows: List[Dict[str, Any]] = []
    previous_progress = _SHOW_FLOW_PROGRESS
    _SHOW_FLOW_PROGRESS = progress is not None
    matrix = api_case_matrix()
    total = len(matrix)
    try:
        with ThreadPoolExecutor(max_workers=min(API_MAX_WORKERS, total)) as executor:
            futures = [executor.submit(run_api_case, case, project_id) for case, project_id in matrix]
            for future in as_completed(futures):
                row = future.result()
                rows.append(row)
                if progress:
                    progress(row, len(rows), total)
    finally:
        _SHOW_FLOW_PROGRESS = previous_progress
    order = {(case.name, project_id): index for index, (case, project_id) in enumerate(matrix)}
    return sorted(rows, key=lambda row: order[(row["case"], row["project"])])


def run_project_api_rows(project_id: str, progress: Optional[Callable[[Dict[str, Any], int, int], None]] = None, total: int = 0) -> List[Dict[str, Any]]:
    flow_progress(f"[project-start] {project_id}")
    started = perf_counter()
    rows: List[Dict[str, Any]] = []
    for case in API_FIXTURE_CHECKS:
        row = run_api_case(case, project_id)
        rows.append(row)
        if progress:
            progress(row, len(rows), total)
    elapsed = perf_counter() - started
    flow_progress(f"[project-done] {project_id} cases={len(rows)} elapsed={elapsed:.1f}s")
    return rows


def write_api_check_csv(output_dir: Optional[Path] = None) -> Path:
    if output_dir is None:
        from datetime import datetime
        output_dir = Path("report/api-check") / datetime.now().strftime("%Y%m%d-%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "api-check.csv"
    fieldnames = [
        "project",
        "case",
        "api",
        "http_status",
        "source",
        "curl_note",
        "curl",
        "curl_script_path",
        "request_json_path",
        "response_json_path",
        "request_body",
        "response_body",
        "expected_schema",
        "checked_response_path",
        "schema_check",
        "schema_error",
    ]
    with csv_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for row in iter_api_project_rows():
            writer.writerow({
                "project": row["project"],
                "case": row["case"],
                "api": row["api"],
                "http_status": row["http_status"],
                "source": row["source"],
                "curl_note": row["curl_note"],
                "curl": row["curl"],
                "curl_script_path": row.get("curl_script_path", ""),
                "request_json_path": row.get("request_json_path", ""),
                "response_json_path": row.get("response_json_path", ""),
                "request_body": json_cell(row["request_body"]),
                "response_body": json_cell(row["response_body"]),
                "expected_schema": row["expected_schema"],
                "checked_response_path": json_cell(row["checked_response_path"]),
                "schema_check": row["schema_check"],
                "schema_error": row["schema_error"],
            })
    return csv_path


def print_progress(row: Dict[str, Any], project_completed: int, total: int) -> None:
    global _PROGRESS_DONE
    with _PROGRESS_LOCK:
        _PROGRESS_DONE += 1
        print(
            f"[{_PROGRESS_DONE}/{total}] {row['project']} {row['case']} "
            f"source={row.get('source', 'api')} http={row['http_status']} schema={row['schema_check']}",
            flush=True,
        )


def write_text_artifact(path: Path, content: str) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")
    return path.as_posix()


def write_json_artifact(path: Path, value: Any) -> str:
    return write_text_artifact(path, pretty_json(value))


def write_curl_artifact(path: Path, command: str) -> str:
    return write_text_artifact(path, f"#!/usr/bin/env bash\nset -euo pipefail\n{command}\n")


def write_api_check_excel(output_dir: Optional[Path] = None, show_progress: bool = False) -> Path:
    global _PROGRESS_DONE
    _PROGRESS_DONE = 0
    if output_dir is None:
        from datetime import datetime
        output_dir = Path("report/api-check") / datetime.now().strftime("%Y%m%d-%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "api-check.xlsx"
    fieldnames = [
        "project",
        "case",
        "api",
        "http_status",
        "source",
        "curl_note",
        "schema_check",
        "schema_error",
        "expected_schema",
        "checked_response_path",
        "curl",
        "curl_script_path",
        "request_json_path",
        "response_json_path",
        "request_body",
        "response_body",
    ]

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "api-check"
    sheet.append(fieldnames)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    rows = iter_api_project_rows(print_progress if show_progress else None)
    for index, row in enumerate(rows, start=1):
        row_id = f"{index:03d}-{row['project']}-{row['case']}".replace("/", "-")
        artifact_dir = output_dir / "rows"
        request_path = write_json_artifact(artifact_dir / f"{row_id}.request.json", row["request_body"])
        response_path = write_json_artifact(artifact_dir / f"{row_id}.response.json", row["response_body"])
        curl_script_path = write_curl_artifact(artifact_dir / f"{row_id}.curl.sh", row["curl"])
        row["curl_script_path"] = curl_script_path
        row["request_json_path"] = request_path
        row["response_json_path"] = response_path
        sheet.append([
            row["project"],
            row["case"],
            row["api"],
            row["http_status"],
            row["source"],
            row["curl_note"],
            row["schema_check"],
            row["schema_error"],
            row["expected_schema"],
            json_cell(row["checked_response_path"]),
            excel_curl_cell(row["curl"], curl_script_path),
            row.get("curl_script_path", ""),
            row.get("request_json_path", ""),
            row.get("response_json_path", ""),
            excel_preview_cell(row["request_body"]),
            excel_preview_cell(row["response_body"]),
        ])

    widths = {
        "A": 24,
        "B": 18,
        "C": 20,
        "D": 12,
        "E": 14,
        "F": 36,
        "G": 42,
        "H": 24,
        "I": 80,
        "J": 80,
        "K": 80,
        "L": 100,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)
    return xlsx_path


def visible_api_report() -> List[Dict[str, Any]]:
    return iter_api_project_rows()


def _check_api_request_with_live_schema(project_id: str, request: Dict[str, Any]) -> bool:
    """挂载 live_schema 校验：api-check 请求体是否符合 REQUEST_SHAPE。校验不阻断。"""
    try:
        import importlib
        ls = importlib.import_module(f"impl.projects.{project_id}.live_schema")
        if hasattr(ls, "check") and isinstance(request, dict):
            return ls.check.request(request)
    except Exception:
        pass
    return True
